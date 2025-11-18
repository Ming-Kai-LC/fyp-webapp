from django.conf import settings
from django.utils import timezone
from django.template.loader import render_to_string
from .models import Report, BatchReportJob
from io import BytesIO
import qrcode
import zipfile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class ReportGenerator:
    """
    Service class for generating PDF reports from predictions
    """
    def __init__(self, prediction, template, include_signature=True,
                 include_logo=True, include_qr=True, custom_notes=''):
        self.prediction = prediction
        self.template = template
        self.include_signature = include_signature
        self.include_logo = include_logo
        self.include_qr = include_qr
        self.custom_notes = custom_notes

    def generate(self, generated_by):
        """
        Generate PDF report and save to database
        """
        # Create Report object
        report = Report.objects.create(
            prediction=self.prediction,
            patient=self.prediction.xray.patient,
            template=self.template,
            generated_by=generated_by,
            include_signature=self.include_signature,
            include_hospital_logo=self.include_logo,
            include_qr_code=self.include_qr,
            custom_notes=self.custom_notes,
            status='generated'
        )

        # Generate QR code if needed
        qr_code_data = None
        if self.include_qr:
            qr_code_data = self._generate_qr_code(report)

        # Prepare context for template
        context = {
            'report': report,
            'prediction': self.prediction,
            'patient': self.prediction.xray.patient,
            'xray': self.prediction.xray,
            'models': self.prediction.get_all_predictions(),
            'best_model': self.prediction.get_best_model(),
            'qr_code': qr_code_data,
            'generated_at': timezone.now(),
        }

        # Render HTML from template
        html_content = render_to_string(
            'reporting/pdf_templates/report_template.html',
            context
        )

        # Convert HTML to PDF using WeasyPrint or similar
        pdf_file = self._html_to_pdf(html_content)

        # Save PDF file
        report.pdf_file.save(
            f'report_{report.report_id}.pdf',
            pdf_file,
            save=True
        )

        report.file_size = report.pdf_file.size
        report.save(update_fields=['file_size'])

        return report

    def _generate_qr_code(self, report):
        """
        Generate QR code for report verification
        """
        # QR code contains report URL or verification code
        verification_url = f"{settings.SITE_URL}/reporting/verify/{report.report_id}/"

        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(verification_url)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # Convert to base64 for embedding in HTML
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)

        import base64
        return base64.b64encode(buffer.getvalue()).decode()

    def _html_to_pdf(self, html_content):
        """
        Convert HTML content to PDF using WeasyPrint
        """
        # Import here to avoid dependency issues if not installed
        try:
            from weasyprint import HTML, CSS
            from django.core.files.base import ContentFile

            # Generate PDF
            pdf_bytes = HTML(string=html_content).write_pdf()

            return ContentFile(pdf_bytes)

        except ImportError:
            # Fallback: use xhtml2pdf or reportlab
            from xhtml2pdf import pisa

            pdf_buffer = BytesIO()
            pisa.CreatePDF(html_content, dest=pdf_buffer)
            pdf_buffer.seek(0)

            return ContentFile(pdf_buffer.getvalue())


class BatchReportProcessor:
    """
    Service class for processing batch report generation jobs
    """
    def __init__(self, batch_job):
        self.batch_job = batch_job

    def process(self):
        """
        Process all predictions in the batch job
        """
        self.batch_job.status = 'processing'
        self.batch_job.save()

        reports = []

        for prediction in self.batch_job.predictions.all():
            try:
                generator = ReportGenerator(
                    prediction=prediction,
                    template=self.batch_job.template,
                )

                report = generator.generate(generated_by=self.batch_job.created_by)
                reports.append(report)

                self.batch_job.completed_reports += 1
                self.batch_job.save(update_fields=['completed_reports'])

            except Exception as e:
                self.batch_job.failed_reports += 1
                self.batch_job.error_log += f"\nFailed for prediction {prediction.id}: {str(e)}"
                self.batch_job.save(update_fields=['failed_reports', 'error_log'])

        # Create ZIP file with all PDFs
        if reports:
            zip_file = self._create_zip(reports)
            self.batch_job.zip_file.save(
                f'batch_{self.batch_job.job_id}.zip',
                zip_file,
                save=True
            )

        self.batch_job.status = 'completed'
        self.batch_job.completed_at = timezone.now()
        self.batch_job.save()

    def _create_zip(self, reports):
        """
        Create ZIP file containing all report PDFs
        """
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for report in reports:
                if report.pdf_file:
                    zip_file.writestr(
                        f'report_{report.report_id}.pdf',
                        report.pdf_file.read()
                    )

        zip_buffer.seek(0)
        from django.core.files.base import ContentFile
        return ContentFile(zip_buffer.getvalue())


class ExcelExporter:
    """
    Service class for exporting prediction data to Excel
    """
    def __init__(self, predictions_queryset):
        self.predictions = predictions_queryset

    def generate(self):
        """
        Generate Excel file with prediction data
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "COVID-19 Predictions"

        # Define headers
        headers = [
            'Report ID', 'Patient Name', 'Patient Age', 'Patient Gender',
            'Upload Date', 'Final Diagnosis', 'Confidence', 'Best Model',
            'CrossViT Prediction', 'CrossViT Confidence',
            'ResNet-50 Prediction', 'ResNet-50 Confidence',
            'DenseNet-121 Prediction', 'DenseNet-121 Confidence',
            'EfficientNet Prediction', 'EfficientNet Confidence',
            'ViT Prediction', 'ViT Confidence',
            'Swin Prediction', 'Swin Confidence',
            'Inference Time (ms)', 'Validated', 'Doctor Notes'
        ]

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Write data
        for row, prediction in enumerate(self.predictions, start=2):
            patient = prediction.xray.patient

            data = [
                str(prediction.id),
                patient.user.get_full_name(),
                patient.age,
                patient.gender,
                prediction.xray.upload_date.strftime('%Y-%m-%d %H:%M'),
                prediction.final_diagnosis,
                f"{prediction.consensus_confidence:.2f}%",
                prediction.get_best_model(),
                prediction.crossvit_prediction,
                f"{prediction.crossvit_confidence:.2f}%",
                prediction.resnet50_prediction,
                f"{prediction.resnet50_confidence:.2f}%",
                prediction.densenet121_prediction,
                f"{prediction.densenet121_confidence:.2f}%",
                prediction.efficientnet_prediction,
                f"{prediction.efficientnet_confidence:.2f}%",
                prediction.vit_prediction,
                f"{prediction.vit_confidence:.2f}%",
                prediction.swin_prediction,
                f"{prediction.swin_confidence:.2f}%",
                f"{prediction.inference_time:.2f}",
                'Yes' if prediction.is_validated else 'No',
                prediction.doctor_notes or 'N/A'
            ]

            for col, value in enumerate(data, start=1):
                sheet.cell(row=row, column=col, value=value)

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer
